from django.conf import settings
from django.utils import timezone
from django.core.files import File
from imbox import Imbox
from openpyxl import load_workbook
from celery import shared_task

from analytics.models import Performance, PerformanceRow


def extract_rows(data):
    workbook = load_workbook(data, read_only=True)
    worksheet = workbook.worksheets[0]
    header = []
    rows = []
    totals = []
    read_header = False

    for row in worksheet.iter_rows():
        if not row[0].value:
            continue

        if not read_header:
            for cell in row:
                header.append(cell.value)
            read_header = not read_header
            continue

        rows.append([cell.value for cell in row])

    totals = rows[-1]
    rows = rows[:-1]

    rows = [dict(zip(header, row)) for row in rows]
    totals = dict(zip(header, totals))


def retrieve_mail_messages():
    hostname = settings.INBOX_HOST
    username = settings.INBOX_USERNAME
    password = settings.INBOX_PASSWORD
    sender = settings.INBOX_SENDER

    msgs = []
    with Imbox(
            hostname,
            username=username,
            password=password,
            ssl=True,
            ssl_context=None,
            starttls=False) as imbox:
        messages = imbox.messages(sent_from=sender)
        for message in messages:
            _, msg = message
            msgs.append(msg)
            p = Performance(received_on=timezone.make_aware(msg.parsed_date, timezone.utc), downloaded_on=timezone.now())
            for attachment in msg.attachments:
                p.worksheet = File(attachment['content'], name=attachment['filename'])
                break
            p.save()
    return msgs


@shared_task
def download_performance_data():
    retrieve_mail_messages()
